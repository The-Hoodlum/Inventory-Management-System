"""Tabular file parsing for .csv / .xlsx / .xls into ``(headers, rows)`` of strings.

openpyxl/xlrd are imported lazily so this module imports even where those wheels
aren't installed (mirrors the lazy-import pattern in the intelligence sources). For
Phase 1 rows are materialized; the streaming hooks for 50k+ land in Phase 3.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass


class UnsupportedFileType(ValueError):
    """Raised for an extension the parser doesn't handle."""


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[str]]


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        # Avoid '100.0' for whole numbers that arrive as floats from Excel.
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, dt.datetime | dt.date):
        return value.isoformat()
    return str(value).strip()


def file_extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def parse_table(filename: str, data: bytes) -> ParsedTable:
    ext = file_extension(filename)
    if ext == "csv":
        return _parse_csv(data)
    if ext == "xlsx":
        return _parse_xlsx(data)
    if ext == "xls":
        return _parse_xls(data)
    raise UnsupportedFileType(f"Unsupported file type: .{ext or '(none)'} (use .csv, .xlsx or .xls)")


def _finalize(matrix: list[list[str]]) -> ParsedTable:
    """First non-empty row = headers; drop fully-blank data rows."""
    matrix = [r for r in matrix if any(c.strip() for c in r)]
    if not matrix:
        return ParsedTable(headers=[], rows=[])
    headers = [c.strip() for c in matrix[0]]
    return ParsedTable(headers=headers, rows=matrix[1:])


def _parse_csv(data: bytes) -> ParsedTable:
    text = data.decode("utf-8-sig", errors="replace")  # tolerate BOM + odd bytes
    reader = csv.reader(io.StringIO(text))
    return _finalize([[_cell_to_str(c) for c in row] for row in reader])


def _parse_xlsx(data: bytes) -> ParsedTable:
    from openpyxl import load_workbook  # lazy

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        matrix = [
            [_cell_to_str(c) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()
    return _finalize(matrix)


def _parse_xls(data: bytes) -> ParsedTable:
    import xlrd  # lazy

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    matrix = [
        [_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]
    return _finalize(matrix)
